#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════╗
║   Buscador de Jurisdicciones CUIT · ARCA                 ║
╚══════════════════════════════════════════════════════════╝
"""

import sys, os, re, json, time, subprocess
from pathlib import Path

# Suprimir errores de Chrome en stderr
import io
sys.stderr = io.StringIO()

def instalar(pkg):
    sys.stderr = sys.__stderr__
    subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"],
                          stderr=subprocess.DEVNULL)
    sys.stderr = io.StringIO()

for pkg in ["selenium","openpyxl","pandas"]:
    try: __import__(pkg)
    except ImportError: instalar(pkg)

import pandas as pd
from selenium import webdriver
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Restaurar stderr para prints propios
sys.stderr = sys.__stderr__

# ════════════════════════════════════════════════════════
# LISTA DE PROVINCIAS VÁLIDAS
# ════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════
# MAPA LOCALIDAD → PROVINCIA
# Incluye capitales, ciudades principales y GBA
# ════════════════════════════════════════════════════════
LOCALIDAD_PROVINCIA = {
    # ── CABA ──
    "ciudad autonoma de buenos aires": "CABA",
    "ciudad autónoma de buenos aires": "CABA",
    "ciudad autonoma buenos aires":    "CABA",
    "ciudad autónoma buenos aires":    "CABA",
    "capital federal":                 "CABA",
    "caba":                            "CABA",
    "buenos aires":                    "CABA",  # cuando dice solo "Buenos Aires" en cuitonline es CABA

    # ── GBA / Buenos Aires ──
    "la plata":             "Buenos Aires",
    "mar del plata":        "Buenos Aires",
    "bahia blanca":         "Buenos Aires",
    "bahía blanca":         "Buenos Aires",
    "quilmes":              "Buenos Aires",
    "lanus":                "Buenos Aires",
    "lanús":                "Buenos Aires",
    "lomas de zamora":      "Buenos Aires",
    "morón":                "Buenos Aires",
    "moron":                "Buenos Aires",
    "san isidro":           "Buenos Aires",
    "vicente lopez":        "Buenos Aires",
    "vicente lópez":        "Buenos Aires",
    "tigre":                "Buenos Aires",
    "pilar":                "Buenos Aires",
    "escobar":              "Buenos Aires",
    "zarate":               "Buenos Aires",
    "zárate":               "Buenos Aires",
    "campana":              "Buenos Aires",
    "san nicolas":          "Buenos Aires",
    "san nicolás":          "Buenos Aires",
    "pergamino":            "Buenos Aires",
    "lujan":                "Buenos Aires",
    "luján":                "Buenos Aires",
    "mercedes":             "Buenos Aires",
    "chivilcoy":            "Buenos Aires",
    "junin":                "Buenos Aires",
    "junín":                "Buenos Aires",
    "azul":                 "Buenos Aires",
    "tandil":               "Buenos Aires",
    "olavarria":            "Buenos Aires",
    "olavarría":            "Buenos Aires",
    "necochea":             "Buenos Aires",
    "tres arroyos":         "Buenos Aires",
    "pehuajo":              "Buenos Aires",
    "pehuajó":              "Buenos Aires",
    "9 de julio":           "Buenos Aires",
    "bragado":              "Buenos Aires",
    "lincoln":              "Buenos Aires",
    "lobos":                "Buenos Aires",
    "cañuelas":             "Buenos Aires",
    "canuelas":             "Buenos Aires",
    "san vicente":          "Buenos Aires",
    "brandsen":             "Buenos Aires",
    "ensenada":             "Buenos Aires",
    "berisso":              "Buenos Aires",
    "la matanza":           "Buenos Aires",
    "merlo":                "Buenos Aires",
    "moreno":               "Buenos Aires",
    "general rodriguez":    "Buenos Aires",
    "general rodríguez":    "Buenos Aires",
    "marcos paz":           "Buenos Aires",
    "navarro":              "Buenos Aires",
    "magdalena":            "Buenos Aires",
    "san pedro":            "Buenos Aires",
    "baradero":             "Buenos Aires",
    "ramallo":              "Buenos Aires",
    "villa constitucion":   "Buenos Aires",
    "villa constitución":   "Buenos Aires",
    "ituzaingo":            "Buenos Aires",
    "ituzaingó":            "Buenos Aires",
    "hurlingham":           "Buenos Aires",
    "tres de febrero":      "Buenos Aires",
    "san martin":           "Buenos Aires",
    "san martín":           "Buenos Aires",
    "jose c paz":           "Buenos Aires",
    "josé c paz":           "Buenos Aires",
    "malvinas argentinas":  "Buenos Aires",
    "florencio varela":     "Buenos Aires",
    "berazategui":          "Buenos Aires",
    "almirante brown":      "Buenos Aires",
    "esteban echeverria":   "Buenos Aires",
    "esteban echeverría":   "Buenos Aires",
    "ezeiza":               "Buenos Aires",
    "presidente peron":     "Buenos Aires",
    "presidente perón":     "Buenos Aires",
    "avellaneda":           "Buenos Aires",
    "san fernando":         "Buenos Aires",
    "general san martin":   "Buenos Aires",
    "general san martín":   "Buenos Aires",
    "el palomar":           "Buenos Aires",
    "haedo":                "Buenos Aires",
    "castelar":             "Buenos Aires",
    "moron":                "Buenos Aires",
    "ramos mejia":          "Buenos Aires",
    "ramos mejía":          "Buenos Aires",
    "ciudadela":            "Buenos Aires",
    "villa liniers":        "Buenos Aires",
    "boulogne":             "Buenos Aires",
    "martinez":             "Buenos Aires",
    "martínez":             "Buenos Aires",
    "olivos":               "Buenos Aires",
    "florida":              "Buenos Aires",
    "munro":                "Buenos Aires",
    "villa del parque":     "Buenos Aires",
    "dont torcuato":        "Buenos Aires",
    "don torcuato":         "Buenos Aires",
    "general pacheco":      "Buenos Aires",
    "jose leon suarez":     "Buenos Aires",
    "jose léon suárez":     "Buenos Aires",
    "gran buenos aires":    "Buenos Aires",
    "gba":                  "Buenos Aires",
    "quilmes oeste":        "Buenos Aires",
    "bernal":               "Buenos Aires",
    "ezpeleta":             "Buenos Aires",
    "wilde":                "Buenos Aires",
    "adrogué":              "Buenos Aires",
    "adrogue":              "Buenos Aires",
    "longchamps":           "Buenos Aires",
    "temperley":            "Buenos Aires",
    "llavallol":            "Buenos Aires",
    "banfield":             "Buenos Aires",
    "remedios de escalada": "Buenos Aires",
    "turdera":              "Buenos Aires",
    "monte grande":         "Buenos Aires",
    "san jose":             "Buenos Aires",
    "canning":              "Buenos Aires",
    "tristan suarez":       "Buenos Aires",
    "tristán suárez":       "Buenos Aires",
    "la union":             "Buenos Aires",
    "pablo nogues":         "Buenos Aires",
    "pablo nogués":         "Buenos Aires",
    "grand bourg":          "Buenos Aires",
    "tortuguitas":          "Buenos Aires",
    "jose ingenieros":      "Buenos Aires",
    "jose ingenieros":      "Buenos Aires",
    "villa adelina":        "Buenos Aires",
    "villa bonich":         "Buenos Aires",
    "san justo":            "Buenos Aires",
    "isidro casanova":      "Buenos Aires",
    "la tablada":           "Buenos Aires",
    "tapiales":             "Buenos Aires",
    "villa luzuriaga":      "Buenos Aires",
    "gregorio de laferrere":"Buenos Aires",
    "rafael castillo":      "Buenos Aires",
    "gonzalez catan":       "Buenos Aires",
    "gonzáles catán":       "Buenos Aires",
    "virrey del pino":      "Buenos Aires",
    "luis guilon":          "Buenos Aires",
    "luis guillón":         "Buenos Aires",
    "luis guillon":         "Buenos Aires",
    "san francisco solano": "Buenos Aires",
    "quilmes":              "Buenos Aires",
    "manuel alberti":       "Buenos Aires",
    "zelaya":               "Buenos Aires",
    "fátima":               "Buenos Aires",
    "fatima":               "Buenos Aires",
    "del viso":             "Buenos Aires",
    "maquinista savio":     "Buenos Aires",
    "ingeniero maschwitz":  "Buenos Aires",
    "garín":                "Buenos Aires",
    "garin":                "Buenos Aires",
    "belén de escobar":     "Buenos Aires",
    "belen de escobar":     "Buenos Aires",
    "open door":            "Buenos Aires",
    "lujan":                "Buenos Aires",
    "villa rosa":           "Buenos Aires",
    "la reja":              "Buenos Aires",
    "paso del rey":         "Buenos Aires",
    "moreno":               "Buenos Aires",
    "francisco álvarez":    "Buenos Aires",
    "francisco alvarez":    "Buenos Aires",
    "san antonio de padua": "Buenos Aires",
    "ituzaingo":            "Buenos Aires",
    "villa sarmiento":      "Buenos Aires",
    "san miguel":           "Buenos Aires",
    "muñiz":                "Buenos Aires",
    "muniz":                "Buenos Aires",
    "campo de mayo":        "Buenos Aires",
    "william morris":       "Buenos Aires",
    "bella vista":          "Buenos Aires",
    "los polvorines":       "Buenos Aires",
    "cuartel v":            "Buenos Aires",
    "trujui":               "Buenos Aires",
    "padua":                "Buenos Aires",
    "merlo":                "Buenos Aires",
    "el talar":             "Buenos Aires",
    "rincon de milberg":    "Buenos Aires",
    "rincón de milberg":    "Buenos Aires",
    "villa la florida":     "Buenos Aires",
    "lomas del mirador":    "Buenos Aires",
    "villa madero":         "Buenos Aires",
    "aldo bonzi":           "Buenos Aires",
    "tapiales":             "Buenos Aires",
    "villa celina":         "Buenos Aires",
    "villa insuperable":    "Buenos Aires",
    "la salada":            "Buenos Aires",
    "ciudad evita":         "Buenos Aires",
    "palermo":              "Buenos Aires",  # aunque es CABA, a veces aparece solo
    "reconquista":          "Santa Fe",  # ojo: hay Reconquista en Santa Fe
    # Corregir Reconquista — es Santa Fe, no confundir
    # Se maneja abajo por orden de prioridad

    # ── Córdoba ──
    "cordoba":              "Córdoba",
    "córdoba":              "Córdoba",
    "rio cuarto":           "Córdoba",
    "río cuarto":           "Córdoba",
    "villa maria":          "Córdoba",
    "villa maría":          "Córdoba",
    "san francisco":        "Córdoba",
    "cosquin":              "Córdoba",
    "cosquín":              "Córdoba",
    "carlos paz":           "Córdoba",
    "villa carlos paz":     "Córdoba",
    "jesus maria":          "Córdoba",
    "jesús maría":          "Córdoba",
    "bell ville":           "Córdoba",
    "rio tercero":          "Córdoba",
    "río tercero":          "Córdoba",
    "alta gracia":          "Córdoba",
    "laboulaye":            "Córdoba",
    "marcos juarez":        "Córdoba",
    "marcos juárez":        "Córdoba",
    "la calera":            "Córdoba",
    "unquillo":             "Córdoba",
    "mendiolaza":           "Córdoba",
    "villa allende":        "Córdoba",
    "rio ceballos":         "Córdoba",
    "río ceballos":         "Córdoba",
    "salsipuedes":          "Córdoba",
    "maipu":                "Córdoba",
    "maipú":                "Córdoba",

    # ── Santa Fe ──
    "santa fe":             "Santa Fe",
    "rosario":              "Santa Fe",
    "rafaela":              "Santa Fe",
    "venado tuerto":        "Santa Fe",
    "reconquista":          "Santa Fe",
    "santo tome":           "Santa Fe",
    "santo tomé":           "Santa Fe",
    "esperanza":            "Santa Fe",
    "villa gobernador galvez":"Santa Fe",
    "villa gobernador gálvez":"Santa Fe",
    "casilda":              "Santa Fe",
    "cañada de gomez":      "Santa Fe",
    "cañada de gómez":      "Santa Fe",
    "firmat":               "Santa Fe",
    "rufino":               "Santa Fe",
    "san lorenzo":          "Santa Fe",
    "granadero baigorria":  "Santa Fe",
    "funes":                "Santa Fe",
    "roldan":               "Santa Fe",
    "roldán":               "Santa Fe",
    "piñero":               "Santa Fe",
    "piñero":               "Santa Fe",
    "armstrong":            "Santa Fe",
    "correa":               "Santa Fe",

    # ── Mendoza ──
    "mendoza":              "Mendoza",
    "godoy cruz":           "Mendoza",
    "guaymallen":           "Mendoza",
    "guaymallén":           "Mendoza",
    "las heras":            "Mendoza",
    "maipu":                "Mendoza",
    "maipú":                "Mendoza",
    "lujan de cuyo":        "Mendoza",
    "luján de cuyo":        "Mendoza",
    "san rafael":           "Mendoza",
    "rivadavia":            "Mendoza",
    "junin":                "Mendoza",  # hay Junín en Mendoza también
    "general alvear":       "Mendoza",
    "malargue":             "Mendoza",
    "malargüe":             "Mendoza",

    # ── Tucumán ──
    "san miguel de tucuman":"Tucumán",
    "san miguel de tucumán":"Tucumán",
    "tucuman":              "Tucumán",
    "tucumán":              "Tucumán",
    "tafi viejo":           "Tucumán",
    "tafí viejo":           "Tucumán",
    "yerba buena":          "Tucumán",
    "banda del rio sali":   "Tucumán",
    "banda del río salí":   "Tucumán",
    "concepcion":           "Tucumán",
    "concepción":           "Tucumán",
    "famailla":             "Tucumán",
    "famaillá":             "Tucumán",

    # ── Salta ──
    "salta":                "Salta",
    "tartagal":             "Salta",
    "oran":                 "Salta",
    "orán":                 "Salta",
    "general guemes":       "Salta",
    "general güemes":       "Salta",
    "cafayate":             "Salta",

    # ── Entre Ríos ──
    "parana":               "Entre Ríos",
    "paraná":               "Entre Ríos",
    "concordia":            "Entre Ríos",
    "gualeguaychu":         "Entre Ríos",
    "gualeguaychú":         "Entre Ríos",
    "gualeguay":            "Entre Ríos",
    "colon":                "Entre Ríos",
    "colón":                "Entre Ríos",
    "villaguay":            "Entre Ríos",
    "victoria":             "Entre Ríos",
    "federacion":           "Entre Ríos",
    "federación":           "Entre Ríos",

    # ── Chaco ──
    "resistencia":          "Chaco",
    "barranqueras":         "Chaco",
    "fontana":              "Chaco",
    "presidencia roque saenz peña": "Chaco",
    "villa angela":         "Chaco",
    "villa ángela":         "Chaco",

    # ── Corrientes ──
    "corrientes":           "Corrientes",
    "goya":                 "Corrientes",
    "mercedes":             "Corrientes",
    "curuzú cuatiá":        "Corrientes",
    "curuzu cuatia":        "Corrientes",
    "paso de los libres":   "Corrientes",
    "santo tome":           "Corrientes",

    # ── Misiones ──
    "posadas":              "Misiones",
    "obera":                "Misiones",
    "oberá":                "Misiones",
    "eldorado":             "Misiones",
    "puerto iguazu":        "Misiones",
    "puerto iguazú":        "Misiones",
    "apostoles":            "Misiones",
    "apóstoles":            "Misiones",

    # ── Santiago del Estero ──
    "santiago del estero":  "Santiago del Estero",
    "la banda":             "Santiago del Estero",
    "termas de rio hondo":  "Santiago del Estero",
    "termas de río hondo":  "Santiago del Estero",
    "frías":                "Santiago del Estero",
    "frias":                "Santiago del Estero",

    # ── San Juan ──
    "san juan":             "San Juan",
    "rivadavia":            "San Juan",
    "rawson":               "San Juan",
    "chimbas":              "San Juan",
    "pocito":               "San Juan",

    # ── Jujuy ──
    "san salvador de jujuy":"Jujuy",
    "jujuy":                "Jujuy",
    "palpalá":              "Jujuy",
    "palpala":              "Jujuy",
    "libertador general san martin": "Jujuy",

    # ── Río Negro ──
    "viedma":               "Río Negro",
    "bariloche":            "Río Negro",
    "san carlos de bariloche": "Río Negro",
    "cipolletti":           "Río Negro",
    "general roca":         "Río Negro",
    "villa regina":         "Río Negro",
    "allen":                "Río Negro",

    # ── Neuquén ──
    "neuquen":              "Neuquén",
    "neuquén":              "Neuquén",
    "plottier":             "Neuquén",
    "centenario":           "Neuquén",
    "cutral co":            "Neuquén",
    "cutral có":            "Neuquén",
    "san martin de los andes": "Neuquén",
    "san martín de los andes": "Neuquén",
    "zapala":               "Neuquén",

    # ── Formosa ──
    "formosa":              "Formosa",
    "clorinda":             "Formosa",

    # ── Chubut ──
    "rawson":               "Chubut",
    "comodoro rivadavia":   "Chubut",
    "trelew":               "Chubut",
    "puerto madryn":        "Chubut",
    "esquel":               "Chubut",

    # ── San Luis ──
    "san luis":             "San Luis",
    "villa mercedes":       "San Luis",
    "merlo":                "San Luis",

    # ── Catamarca ──
    "san fernando del valle de catamarca": "Catamarca",
    "catamarca":            "Catamarca",

    # ── La Rioja ──
    "la rioja":             "La Rioja",
    "chilecito":            "La Rioja",

    # ── La Pampa ──
    "santa rosa":           "La Pampa",
    "general pico":         "La Pampa",

    # ── Santa Cruz ──
    "rio gallegos":         "Santa Cruz",
    "río gallegos":         "Santa Cruz",
    "caleta olivia":        "Santa Cruz",
    "pico truncado":        "Santa Cruz",
    "las heras":            "Santa Cruz",

    # ── Tierra del Fuego ──
    "ushuaia":              "Tierra del Fuego",
    "rio grande":           "Tierra del Fuego",
    "río grande":           "Tierra del Fuego",
}

def localidad_a_provincia(texto):
    """Convierte localidad a provincia. Busca la coincidencia más larga."""
    if not texto: return None
    t = texto.strip().lower()
    # Normalizar caracteres
    t = t.replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u').replace('ü','u')
    mejor = None
    largo = 0
    for k, v in LOCALIDAD_PROVINCIA.items():
        k_norm = k.replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u').replace('ü','u')
        if k_norm == t or k_norm in t:
            if len(k_norm) > largo:
                mejor = v
                largo = len(k_norm)
    return mejor

def es_provincia_valida(texto):
    return localidad_a_provincia(texto)

# ════════════════════════════════════════════════════════
# HELPERS
# ════════════════════════════════════════════════════════
def fmt_cuit(raw):
    s = re.sub(r'\D','',str(raw)).zfill(11)
    return f"{s[:2]}-{s[2:10]}-{s[10]}" if len(s)>=11 else s

def raw_cuit(c): return re.sub(r'\D','',str(c))

def tipo_persona(cuit):
    p = raw_cuit(cuit)[:2]
    if p in ["20","21","23","24","27"]: return "Física"
    if p in ["30","33","34"]: return "Jurídica"
    return "Otra"

# ════════════════════════════════════════════════════════
# CACHE
# ════════════════════════════════════════════════════════
CACHE_FILE = "cache_cuits.json"

def cargar_cache():
    if Path(CACHE_FILE).exists():
        try: return json.load(open(CACHE_FILE,"r",encoding="utf-8"))
        except: pass
    return {}

def guardar_cache(cache):
    json.dump(cache, open(CACHE_FILE,"w",encoding="utf-8"), ensure_ascii=False, indent=2)

# ════════════════════════════════════════════════════════
# EDGE — modo visible pero minimizado
# ════════════════════════════════════════════════════════
def buscar_driver():
    for base in [
        r"C:\Program Files (x86)\Microsoft\Edge\Application",
        r"C:\Program Files\Microsoft\Edge\Application",
    ]:
        p = Path(base)
        if p.exists():
            for sub in p.iterdir():
                d = sub / "msedgedriver.exe"
                if d.exists(): return str(d)
    local = Path(__file__).parent / "msedgedriver.exe"
    if local.exists(): return str(local)
    return None

def iniciar_edge():
    opts = EdgeOptions()
    opts.add_argument("--window-position=-2000,0")
    opts.add_argument("--window-size=1280,900")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--log-level=3")
    opts.add_argument("--silent")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-logging", "enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/136.0.0.0 Safari/537.36 Edg/136.0.0.0")

    try:
        service = EdgeService(log_output=subprocess.DEVNULL)
        return webdriver.Edge(service=service, options=opts)
    except:
        pass

    dp = buscar_driver()
    if dp:
        try:
            service = EdgeService(executable_path=dp, log_output=subprocess.DEVNULL)
            return webdriver.Edge(service=service, options=opts)
        except: pass

    print("  No se pudo iniciar Edge.")
    input("  Presiona Enter para cerrar...")
    sys.exit(1)

# ════════════════════════════════════════════════════════
# SCRAPING
# ════════════════════════════════════════════════════════
NORM_PROV = {
    'Buenos Aires': 'Buenos Aires', 'BUENOS AIRES': 'Buenos Aires',
    'Córdoba': 'Córdoba', 'CORDOBA': 'Córdoba', 'Cordoba': 'Córdoba',
    'Santa Fe': 'Santa Fe', 'SANTA FE': 'Santa Fe',
    'Mendoza': 'Mendoza', 'MENDOZA': 'Mendoza',
    'Tucumán': 'Tucumán', 'TUCUMAN': 'Tucumán', 'Tucuman': 'Tucumán',
    'Salta': 'Salta', 'ENTRE RIOS': 'Entre Ríos', 'Entre Rios': 'Entre Ríos',
    'Entre Ríos': 'Entre Ríos', 'Chaco': 'Chaco', 'Corrientes': 'Corrientes',
    'Misiones': 'Misiones', 'Santiago del Estero': 'Santiago del Estero',
    'SANTIAGO DEL ESTERO': 'Santiago del Estero', 'San Juan': 'San Juan',
    'Jujuy': 'Jujuy', 'Río Negro': 'Río Negro', 'RIO NEGRO': 'Río Negro',
    'Rio Negro': 'Río Negro', 'Neuquén': 'Neuquén', 'NEUQUEN': 'Neuquén',
    'Neuquen': 'Neuquén', 'Formosa': 'Formosa', 'Chubut': 'Chubut',
    'San Luis': 'San Luis', 'Catamarca': 'Catamarca', 'La Rioja': 'La Rioja',
    'La Pampa': 'La Pampa', 'Santa Cruz': 'Santa Cruz',
    'Tierra del Fuego': 'Tierra del Fuego', 'TIERRA DEL FUEGO': 'Tierra del Fuego',
    'CABA': 'CABA', 'Capital Federal': 'CABA', 'CAPITAL FEDERAL': 'CABA',
    'Ciudad Autónoma de Buenos Aires': 'CABA', 'Ciudad Autonoma de Buenos Aires': 'CABA',
    'Buenos Aires (Pcia)': 'Buenos Aires', 'Salta': 'Salta', 'SALTA': 'Salta',
    'CHACO': 'Chaco', 'CORRIENTES': 'Corrientes', 'MISIONES': 'Misiones',
    'SAN JUAN': 'San Juan', 'JUJUY': 'Jujuy', 'FORMOSA': 'Formosa',
    'CHUBUT': 'Chubut', 'SAN LUIS': 'San Luis', 'CATAMARCA': 'Catamarca',
    'LA RIOJA': 'La Rioja', 'LA PAMPA': 'La Pampa', 'SANTA CRUZ': 'Santa Cruz',
    'MENDOZA': 'Mendoza', 'SANTA FE': 'Santa Fe', 'CORDOBA': 'Córdoba',
    'TUCUMAN': 'Tucumán',
}

def norm(r):
    return NORM_PROV.get(r.strip(), r.strip()) if r else None

def scrape_cuitonline(driver, cuit_fmt):
    """Scraping de cuitonline.com — addressRegion + addressLocality."""
    raw = raw_cuit(cuit_fmt)
    try:
        driver.get(f"https://www.cuitonline.com/detalle/{raw}/")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        time.sleep(2.0)
        region = None
        locality = None
        try:
            region = driver.find_element(By.CSS_SELECTOR, '[itemprop="addressRegion"]').text.strip() or None
        except: pass
        try:
            locality = driver.find_element(By.CSS_SELECTOR, '[itemprop="addressLocality"]').text.strip() or None
        except: pass
        if region:
            if region.strip().lower() == 'buenos aires' and locality:
                if any(x in locality.lower() for x in ['ciudad autonoma', 'ciudad autónoma', 'capital federal']):
                    return 'CABA'
            return norm(region)
        if locality:
            if any(x in locality.lower() for x in ['ciudad autonoma', 'ciudad autónoma', 'capital federal']):
                return 'CABA'
            p = es_provincia_valida(locality)
            if p: return p
        for linea in driver.find_element(By.TAG_NAME, "body").text.split('\n'):
            if re.match(r'\s*Provincia\s*:', linea, re.I):
                valor = re.sub(r'^.*?Provincia\s*:\s*', '', linea, flags=re.I).split('-')[0].strip()
                if valor: return norm(valor) or es_provincia_valida(valor)
    except: pass
    return None

def scrape_dateas(driver, cuit_fmt):
    """Scraping de dateas.com — tabla col3."""
    from selenium.webdriver.common.keys import Keys
    raw = raw_cuit(cuit_fmt)
    try:
        driver.get('https://www.dateas.com/es/consulta_cuit_cuil')
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'cuit-cuil-dni')))
        inp = driver.find_element(By.ID, 'cuit-cuil-dni')
        inp.clear()
        inp.send_keys(raw)
        inp.send_keys(Keys.RETURN)
        time.sleep(3)
        for fila in driver.find_elements(By.CSS_SELECTOR, 'table tr')[1:]:
            celdas = fila.find_elements(By.TAG_NAME, 'td')
            if len(celdas) >= 4:
                if celdas[1].text.strip().replace('-','').replace(' ','') == raw:
                    prov = celdas[3].text.strip()
                    if prov: return norm(prov)
    except: pass
    return None


# ════════════════════════════════════════════════════════
# ARCHIVO
# ════════════════════════════════════════════════════════
def detectar_archivo():
    archivos = [p for ext in ["*.csv","*.xlsx","*.xls"]
                for p in Path(".").glob(ext)
                if not p.name.startswith(("jurisdicciones_","~$","cache_"))]
    if not archivos:
        print("\n❌ No se encontró CSV o Excel en esta carpeta.")
        input(); sys.exit(1)
    if len(archivos) == 1: return archivos[0]
    print("\n📂 Archivos encontrados:")
    for i,a in enumerate(archivos): print(f"   {i+1}. {a.name}")
    while True:
        try:
            n = int(input("\nElegí el número: ")) - 1
            if 0 <= n < len(archivos): return archivos[n]
        except: pass

def leer_archivo(path):
    if path.suffix.lower() == ".csv":
        for enc in ["windows-1252","utf-8","latin-1"]:
            try:
                with open(path,"r",encoding=enc) as f: muestra = f.read(2000)
                sep = ";" if muestra.count(";") > muestra.count(",") else ","
                return pd.read_csv(path, sep=sep, encoding=enc, dtype=str, low_memory=False)
            except: continue
    return pd.read_excel(path, dtype=str)

def detectar_columnas(df):
    cols = list(df.columns)
    cuit_col = next((c for c in cols if re.search(r'nro.*doc.*emisor',c,re.I)), None) \
            or next((c for c in cols if re.search(r'nro.*doc',c,re.I)
                     and re.search(r'emisor',c,re.I)), None) \
            or (cols[7] if len(cols)>7 else cols[0])
    name_col = next((c for c in cols if re.search(r'denominaci.*emisor',c,re.I)), None) \
            or next((c for c in cols if re.search(r'denominaci|razon.*social',c,re.I)), None) \
            or (cols[8] if len(cols)>8 else cols[1])
    return cuit_col, name_col

def extraer_cuits(df, cuit_col, name_col):
    cuits = {}
    for _, row in df.iterrows():
        raw = re.sub(r'\D','',str(row.get(cuit_col,"") or ""))
        if len(raw) < 10: continue
        fmt = fmt_cuit(raw)
        if fmt not in cuits:
            cuits[fmt] = str(row.get(name_col,"") or "").strip() or "—"
    return cuits

# ════════════════════════════════════════════════════════
# EXPORTAR
# ════════════════════════════════════════════════════════
def exportar_excel(resultados, stem, df_original, cuit_col):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from copy import copy

    # ── Agregar columna Provincia al DataFrame original ──
    df_out = df_original.copy()

    # Encontrar índice de la columna CUIT emisor
    cols = list(df_out.columns)
    cuit_idx = cols.index(cuit_col)

    # Construir serie de provincia mapeando por CUIT
    def get_prov(row):
        raw = re.sub(r'\D','',str(row.get(cuit_col,"") or ""))
        if len(raw) < 10: return ""
        fmt = fmt_cuit(raw)
        return resultados.get(fmt, {}).get("provincia") or ""

    df_out["__provincia__"] = df_out.apply(get_prov, axis=1)

    # Insertar columna después de cuit_col
    insert_pos = cuit_idx + 1
    cols_new = cols[:insert_pos] + ["Provincia"] + cols[insert_pos:]
    df_out.rename(columns={"__provincia__": "Provincia"}, inplace=True)
    df_out = df_out[cols_new]

    salida = f"jurisdicciones_{stem}.xlsx"

    with pd.ExcelWriter(salida, engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False, sheet_name="Ventas con Provincia")
        ws = writer.sheets["Ventas con Provincia"]

        # Encontrar la columna Provincia en el sheet
        prov_col_idx = list(df_out.columns).index("Provincia") + 1  # 1-based
        prov_col_letter = get_column_letter(prov_col_idx)

        # ── Estilo encabezados ──
        header_fill    = PatternFill("solid", fgColor="1E3A5F")   # azul oscuro
        header_font    = Font(color="FFFFFF", bold=True, size=10)
        prov_hdr_fill  = PatternFill("solid", fgColor="1A6B3C")   # verde oscuro
        prov_hdr_font  = Font(color="FFFFFF", bold=True, size=10)
        center         = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for cell in ws[1]:
            if cell.column == prov_col_idx:
                cell.fill = prov_hdr_fill
                cell.font = prov_hdr_font
            else:
                cell.fill = header_fill
                cell.font = header_font
            cell.alignment = center

        # ── Estilo filas de datos ──
        fill_prov_ok   = PatternFill("solid", fgColor="E8F5E9")   # verde claro
        fill_prov_miss = PatternFill("solid", fgColor="FFEBEE")   # rojo claro
        fill_alt       = PatternFill("solid", fgColor="F5F8FF")   # gris azulado
        fill_white     = PatternFill("solid", fgColor="FFFFFF")
        font_prov      = Font(color="1A6B3C", bold=True, size=10)
        font_prov_miss = Font(color="C62828", bold=True, size=10)
        font_normal    = Font(size=10)
        thin_side      = Side(style="thin", color="D0D7E3")
        thin_border    = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill_row = fill_alt if row_idx % 2 == 0 else fill_white
            for cell in row:
                cell.border = thin_border
                cell.font = font_normal
                cell.alignment = Alignment(vertical="center")
                if cell.column == prov_col_idx:
                    val = cell.value or ""
                    if val:
                        cell.fill = fill_prov_ok
                        cell.font = font_prov
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.fill = fill_prov_miss
                        cell.font = font_prov_miss
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.fill = fill_row

        # ── Ancho de columnas ──
        for col_idx, col_name in enumerate(df_out.columns, start=1):
            letter = get_column_letter(col_idx)
            if col_name == "Provincia":
                ws.column_dimensions[letter].width = 22
            else:
                # Auto-ajuste básico
                max_len = max(
                    len(str(col_name)),
                    max((len(str(ws.cell(r, col_idx).value or "")) for r in range(2, min(50, ws.max_row+1))), default=0)
                )
                ws.column_dimensions[letter].width = min(max(max_len + 2, 8), 40)

        # ── Altura de encabezado ──
        ws.row_dimensions[1].height = 32

        # ── Freeze panes ──
        ws.freeze_panes = "A2"

        # ── Filtros ──
        ws.auto_filter.ref = ws.dimensions

    return salida

# ════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════
def main():
    print("\n" + "═"*62)
    print("  🏛️  Buscador de Jurisdicciones CUIT · ARCA")
    print("═"*62)

    archivo = detectar_archivo()
    print(f"\n  ✅ Archivo: {archivo.name}")
    df = leer_archivo(archivo)
    cuit_col, name_col = detectar_columnas(df)
    cuits = extraer_cuits(df, cuit_col, name_col)
    total = len(cuits)

    cache = cargar_cache()
    desde_cache = sum(1 for c in cuits if c in cache)
    a_consultar = total - desde_cache

    print(f"     {len(df):,} filas  |  {total} CUITs únicos  |  {desde_cache} en caché  |  {a_consultar} a consultar")

    if a_consultar == 0:
        print("\n  ✅ Todos en caché. Generando Excel...")
    else:
        t = int(a_consultar * 2.0)  # ~2 workers en paralelo
        print(f"  ⏱  Tiempo estimado: ~{t//60}m {t%60}s")
        print("  (Se abren 2 ventanas de Edge minimizadas — no las cierres)")
        print("  (Ctrl+C para detener y guardar lo procesado)")
        input("\n  Presioná Enter para comenzar...")

    resultados = {}
    ok = 0
    sin_datos = 0

    # Cargar resultados desde caché
    for fmt, nombre in cuits.items():
        if fmt in cache:
            prov = cache[fmt].get("provincia")
            resultados[fmt] = {"nombre":nombre,"provincia":prov,"estado":"Desde caché"}
            if prov: ok += 1

    pendientes = [(fmt, nombre) for fmt, nombre in cuits.items() if fmt not in cache]

    if pendientes:
        print(f"\n  {'N°':>4}  {'CUIT':<16}  {'Provincia':<26}  Razón Social")
        print(f"  {'─'*4}  {'─'*16}  {'─'*26}  {'─'*30}")
        procesados_lock = __import__('threading').Lock()
        contador = [0]

        import random, threading

        print("  Iniciando 2 navegadores (cuitonline + dateas)...\n")

        # Dividir pendientes en 2 lotes alternados
        lote_cuitonline = pendientes[0::2]
        lote_dateas     = pendientes[1::2]

        drivers = []
        lock = threading.Lock()

        def worker(lote, fuente):
            drv = iniciar_edge()
            with lock:
                drivers.append(drv)
            try:
                for fmt, nombre in lote:
                    prov = None
                    # Intentar con el sitio asignado
                    if fuente == 'cuitonline':
                        prov = scrape_cuitonline(drv, fmt)
                    else:
                        prov = scrape_dateas(drv, fmt)

                    # Si falla, intentar con el otro sitio
                    if not prov:
                        if fuente == 'cuitonline':
                            prov = scrape_dateas(drv, fmt)
                        else:
                            prov = scrape_cuitonline(drv, fmt)

                    with lock:
                        if prov:
                            cache[fmt] = {"provincia": prov}
                            guardar_cache(cache)
                            estado = "Encontrado"
                            icono = "✅"
                        else:
                            estado = "Sin datos"
                            icono = "❌"
                        resultados[fmt] = {"nombre":nombre,"provincia":prov,"estado":estado}
                        contador[0] += 1
                        nombre_corto = (nombre[:28] + "..") if len(nombre) > 30 else nombre
                        prov_mostrar = prov or "Sin datos"
                        sitio = "C" if fuente == "cuitonline" else "D"
                        print(f"  {contador[0]:>4} [{sitio}]  {fmt:<16}  {icono} {prov_mostrar:<24}  {nombre_corto}")

                    time.sleep(random.uniform(1.2, 2.0))
            finally:
                try: drv.quit()
                except: pass

        t1 = threading.Thread(target=worker, args=(lote_cuitonline, 'cuitonline'), daemon=True)
        t2 = threading.Thread(target=worker, args=(lote_dateas, 'dateas'), daemon=True)

        try:
            t1.start()
            time.sleep(2)  # escalonar inicio
            t2.start()
            t1.join()
            t2.join()
        except KeyboardInterrupt:
            print("\n\n  Detenido. Guardando resultados parciales...")
            for fmt, nombre in cuits.items():
                if fmt not in resultados:
                    resultados[fmt] = {"nombre":nombre,"provincia":None,"estado":"Pendiente"}
        finally:
            print("  Cerrando navegadores...", end=" ", flush=True)
            for drv in drivers:
                try: drv.quit()
                except: pass
            print("OK")

        ok = sum(1 for r in resultados.values() if r.get("provincia"))
        sin_datos = sum(1 for r in resultados.values() if not r.get("provincia") and r.get("estado") != "Desde caché")

    total_ok = ok + desde_cache
    print(f"\n  {'═'*60}")
    print(f"  ✅ Con provincia: {total_ok}/{total}")
    print(f"  ❌ Sin datos:     {sin_datos}/{total}")
    print(f"  💾 Desde caché:   {desde_cache}/{total}")

    salida = exportar_excel(resultados, archivo.stem, df, cuit_col)
    print(f"\n  📥 Excel: {salida}")
    print(f"  📁 {Path('.').resolve()}")
    print("═"*62)
    input("\n  Presioná Enter para cerrar...")

if __name__ == "__main__":
    main()
